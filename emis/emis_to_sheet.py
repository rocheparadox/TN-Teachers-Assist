# Author : Roche Christopher
# 20/05/22 - 07:58:27

import json
import os
from openpyxl import Workbook
import datetime
import requests
from argparse import ArgumentParser


# change it as per convenience
emis_code_to_header = {
    'unique_id_no': 'EMIS no',
    'name': 'Name',
    'dob': 'DOB',
    'aadhaar_uid_number': 'Aadhaar number',
    'father_name': "Father's Name",
    'mother_name': "Mother's name",
    'phone_number': "Phone No",
    'group': 'Blood Group',
    'religion_name': 'Religion',
    'community_id': 'Community',
    'ifsc_code': "IFSC",
    'account_no': "Account No"
}


def _get_caste_code_from_emis_server() -> dict:
    try:
        print('Trying to fetch caste data from emis server...')
        res = requests.get('https://emis.tnschools.gov.in/assets/json/caste.json')
        caste_json = json.loads(res.text)
    except Exception as exc:
        print("An exception occurred while trying to get the caste data from server. "
              "So using the data from caste.json file")
        with open('caste.json') as castes:
            caste_json=json.loads(castes.read())
    return caste_json['caste']


def _get_community_code_from_emis_server() -> dict:
    try:
        print('Trying to fetch community data from emis server...')
        res = requests.get('https://emis.tnschools.gov.in/assets/json/community.json')
        community_json = json.loads(res.text)
    except Exception as exc:
        print("An error occurred while trying to get the community data from the server. "
              "So using the data from communitu.json file")
        with open('community.json') as communities:
            community_json = json.loads(communities.read())

    return community_json['community']


def _get_static_data_from_emis_server():
    global communities, castes
    communities = _get_community_code_from_emis_server()
    castes = _get_caste_code_from_emis_server()


def _convert_emis_data_to_json(file_path) -> dict:
    if not os.path.exists(file_path):
        print('file does not exist')
        exit(1)

    with open(file_path, 'r') as file_content:
        try:
            json_data = json.loads(file_content.read())
            return json_data['result']
        except Exception as exc:
            print(f"An error occurred while trying to convert the contents of {file_path} to json", exc)


def _store_data_to_sheet(result):
    wb = Workbook()
    ws = wb.active
    students = []

    # write headers to excel sheet
    for col, emis_code in enumerate(emis_code_to_header, start=1):
        ws.cell(row=1, column=col).value = emis_code_to_header[emis_code]

    # row set to 2 since first row is written with headers
    row = 2

    for student in result:
        details = []
        for emis_code in emis_code_to_header:
            # handle the modifications here

            if emis_code == 'dob':
                details.append(datetime.datetime.strptime(student['dob'], '%Y-%M-%d').strftime('%d-%M-%Y'))
                continue

            if emis_code == 'community_id':
                for community in communities:
                    if str(community['id']) == str(student[emis_code]):
                        details.append(community['community_name'])
                        break
                continue

            details.append(student[emis_code])

        # update data to the list
        print(details)
        students.append(details)
        for col, val in enumerate(details, start=1):
            ws.cell(row=row, column=col).value = val
        row += 1  # increment the row, so the next student's data can be written in the next row

    print('Saving data to sheet')
    # save the sheet
    wb.save(output_file_path)
    print(f'data saved to sheet {output_file_path}')


if __name__ == '__main__':
    my_parser = ArgumentParser()
    my_parser.add_argument('--input',
                           required=True,
                           help='path of the input file containing the json data of the students')
    my_parser.add_argument('--output',
                           help='path of the output file to which the xlsx data has to be written')
    args = my_parser.parse_args()
    input_file_path = args.input
    output_file_path = args.output

    if output_file_path is None:
        output_file_path: str = input_file_path
        if output_file_path.endswith(".json"):
            output_file_path = output_file_path.replace('.json', '.xlsx')
        if output_file_path.endswith(".txt"):
            output_file_path = output_file_path.replace('.txt', '.xlsx')

        print(f'output file will be written to {output_file_path}')

    if not os.path.exists(input_file_path):
        print('input file path does not exist')
        exit(1)

    communities = {}
    castes = {}
    _get_static_data_from_emis_server()
    result = _convert_emis_data_to_json(input_file_path)
    _store_data_to_sheet(result)






